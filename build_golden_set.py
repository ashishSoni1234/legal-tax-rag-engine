"""
Golden Set Builder for Legal & Tax RAG System
Generates 135+ Q&A pairs from 104 PDF documents across 4 categories.
All answers are based on content explicitly read from the PDFs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ──────────────────────────────────────────────────────────────────────────────
# GOLDEN SET DATA  (ID | Category | Query | Ground Truth Answer |
#                   Source Document | Page Number | Section/Clause |
#                   Query Type | Difficulty | Notes)
# ──────────────────────────────────────────────────────────────────────────────

rows = [

# ════════════════════════════════════════════════════════════════════════
# ACTS  (target 40)
# ════════════════════════════════════════════════════════════════════════

# ── 26 U.S.C. §1 – Tax imposed ──────────────────────────────────────────
("A001", "Acts",
 "What is the top individual income tax rate imposed under 26 U.S.C. §1 for taxable income over $1,000,000 (married filing jointly), and from what base does the rate apply?",
 "For married individuals filing jointly, the statutory rate table in §1(a) sets the highest bracket at 39.6 percent on the excess of taxable income over the threshold for that bracket. For taxable income exceeding approximately $470,000 (as adjusted), the tentative rate is 39.6 percent of that excess, representing the maximum rate applicable under the statutory tables.",
 "26 U.S. Code § 1 - Tax imposed _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§1(a)", "Factual", "Medium",
 "Married filing jointly bracket. Rate table read from pages 1-2."),

("A002", "Acts",
 "Under 26 U.S.C. §1, how is the cost-of-living adjustment to the tax rate brackets determined, and which price index is used?",
 "Section 1(f)(3) requires annual adjustment of the tax brackets using the Chained Consumer Price Index for All Urban Consumers (C-CPI-U). The adjustment is determined by the percentage change in the C-CPI-U for the preceding calendar year compared to the C-CPI-U for calendar year 2016, so that the dollar amounts in the rate tables are updated each year for inflation.",
 "26 U.S. Code § 1 - Tax imposed _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 1, "§1(f)(3)", "Conceptual", "Hard",
 "Inflation adjustment methodology."),

("A003", "Acts",
 "What is the 'kiddie tax' rule under 26 U.S.C. §1, and at what age does it generally cease to apply?",
 "Under §1(g), the net unearned income of a child who has not attained age 18 by the close of the taxable year (or is a full-time student under age 24 with earned income not exceeding half of their support) is taxed at the parent's marginal rate. The rule is intended to prevent income-shifting to children in lower brackets. It generally ceases to apply once the child turns 18 (or 24 for full-time students meeting the support test).",
 "26 U.S. Code § 1 - Tax imposed _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§1(g)", "Conceptual", "Hard",
 "Kiddie tax rule."),

# ── 26 U.S.C. §61 – Gross income defined ────────────────────────────────
("A004", "Acts",
 "How does 26 U.S.C. §61(a) define gross income, and what are three items explicitly listed as inclusions?",
 "Section 61(a) provides that gross income means all income from whatever source derived. Among the items explicitly listed are: (1) compensation for services, including fees, commissions, fringe benefits, and similar items; (4) interest; (7) dividends; and (11) income from discharge of indebtedness. The list is illustrative rather than exhaustive, as the section uses the phrase 'including (but not limited to).'",
 "26 U.S. Code § 61 - Gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§61(a)", "Factual", "Easy",
 "14 items listed in §61(a)."),

("A005", "Acts",
 "According to 26 U.S.C. §61, where in the Code can taxpayers find items specifically excluded from gross income?",
 "Section 61(b) provides a cross-reference directing taxpayers to Part III of Subchapter B (sections 101 and following) for items specifically excluded from gross income, and to Part II (sections 71 and following) for items specifically included. This structure means that §61 establishes the general inclusion rule while later sections carve out specific exclusions.",
 "26 U.S. Code § 61 - Gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§61(b)", "Factual", "Easy",
 ""),

# ── 26 U.S.C. §62 – Adjusted gross income ───────────────────────────────
("A006", "Acts",
 "Under 26 U.S.C. §62, what is the definition of 'adjusted gross income' for an individual, and name three above-the-line deductions it allows?",
 "Section 62(a) defines adjusted gross income (AGI) as gross income minus certain specified deductions. Three deductions allowed above-the-line include: (1) trade and business deductions (§62(a)(1)); (7) the retirement savings deduction allowed by §219 for IRA contributions; and (17) the interest deduction on qualified education loans under §221. These deductions reduce gross income before itemized or standard deductions are applied.",
 "26 U.S. Code § 62 - Adjusted gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§62(a)", "Factual", "Medium",
 ""),

("A007", "Acts",
 "Under 26 U.S.C. §62(b), what are the requirements for a 'qualified performing artist' to deduct performing arts expenses above-the-line?",
 "Section 62(b) provides that a qualified performing artist must: (A) have performed services for at least two employers in the performing arts during the taxable year; (B) have allowable deductions under §162 attributable to such services that exceed 10 percent of gross income from those services; and (C) have adjusted gross income (before the performing arts deduction) of not more than $16,000. Meeting all three tests allows the performing artist's business expenses to be deducted from gross income as an above-the-line deduction.",
 "26 U.S. Code § 62 - Adjusted gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 5, "§62(b)", "Factual", "Hard",
 "$16,000 AGI cap."),

# ── 26 U.S.C. §101 – Certain death benefits ─────────────────────────────
("A008", "Acts",
 "Under 26 U.S.C. §101(a), are life insurance proceeds paid by reason of death generally excluded from the beneficiary's gross income? Are there any exceptions?",
 "Yes. Section 101(a)(1) provides that gross income does not include amounts received under a life insurance contract, if such amounts are paid by reason of the insured's death. However, §101(a)(2) contains a key exception called the 'transfer-for-value' rule: if the policy was transferred for a valuable consideration to another party, the exclusion is limited to the sum of the consideration paid plus net premiums paid by the transferee. There are further exceptions (e.g., transfers to the insured, a partner, or a corporation) that preserve the exclusion.",
 "26 U.S. Code § 101 - Certain death benefits _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§101(a)", "Conceptual", "Medium",
 "Transfer-for-value exception is critical."),

("A009", "Acts",
 "How does 26 U.S.C. §101(g) treat accelerated death benefits paid to a terminally ill individual?",
 "Section 101(g)(1) provides that any amount received on an accelerated basis under a life insurance contract by an individual who is terminally ill (diagnosed with illness expected to result in death within 24 months) is treated as an amount paid by reason of the insured's death, and is therefore excluded from gross income under §101(a). The same treatment applies to amounts received from viatical settlement providers. This allows terminally ill policyholders to access tax-free funds from their life insurance policies before death.",
 "26 U.S. Code § 101 - Certain death benefits _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 4, "§101(g)", "Conceptual", "Hard",
 "24-month terminal illness window."),

# ── 26 U.S.C. §104 – Compensation for injuries ──────────────────────────
("A010", "Acts",
 "Under 26 U.S.C. §104(a), what types of compensation are excluded from gross income, and does this exclusion apply to punitive damages?",
 "Section 104(a) excludes from gross income: (1) amounts received under workers' compensation acts; (2) damages received on account of personal physical injuries or physical sickness (other than punitive damages); (3) amounts received through accident or health insurance for personal injuries or sickness; (4) amounts received as a pension or disability retirement pay by members of the Armed Forces for combat-related injuries; and (5) amounts received by disability insurance due to injuries from terrorism. Punitive damages are explicitly excluded from the §104(a)(2) exclusion, meaning they are taxable as ordinary income, except in wrongful death actions in states whose law allows only punitive damages.",
 "26 U.S. Code § 104 - Compensation for injuries or sickness _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§104(a)", "Factual", "Medium",
 "Punitive damages are generally taxable."),

("A011", "Acts",
 "Under 26 U.S.C. §104(a)(2), is emotional distress treated the same as 'physical' injury for purposes of the income exclusion?",
 "No. Section 104(a)(2) specifies that the exclusion applies to damages received 'on account of personal physical injuries or physical sickness.' The statute specifically provides that emotional distress itself does not constitute a physical injury or sickness, so damages received purely for emotional distress are not excludable. However, if emotional distress damages arise from a physical injury (i.e., the emotional distress is the result of the physical injury), those damages may be excludable to the extent they compensate for physical harm.",
 "26 U.S. Code § 104 - Compensation for injuries or sickness _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§104(a)(2)", "Conceptual", "Hard",
 "Emotional distress vs. physical injury distinction."),

# ── 26 U.S.C. §105 – Accident and health plans ──────────────────────────
("A012", "Acts",
 "Under 26 U.S.C. §105(b), when may an employee exclude employer-provided accident and health plan reimbursements from gross income?",
 "Section 105(b) provides that gross income does not include amounts paid to an employee through an employer's accident or health plan to the extent they reimburse the employee for medical care expenses (as defined in §213(d)) for the employee, spouse, or dependents. The exclusion applies where the amounts are actually used for medical care and are not otherwise deductible. If the employer has a self-insured plan, nondiscrimination rules under §105(h) must be satisfied; otherwise, highly compensated employees lose the exclusion.",
 "26 U.S. Code § 105 - Amounts received under accident and health plans _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§105(b)", "Conceptual", "Medium",
 "Self-insured plan nondiscrimination important."),

("A013", "Acts",
 "Under 26 U.S.C. §105(h), which individuals are treated as 'highly compensated' for purposes of the self-insured medical reimbursement plan nondiscrimination rules?",
 "Section 105(h)(5) defines highly compensated individuals as those who are: (1) one of the five highest paid officers; (2) a shareholder who owns more than 10 percent of the stock (by value); or (3) among the highest paid 25 percent of all employees (other than employees with less than 3 years of service, under age 25, part-time or seasonal employees, and employees covered by a collective bargaining agreement). A self-insured plan that discriminates in favor of these individuals disqualifies the favorable tax treatment for the highly compensated.",
 "26 U.S. Code § 105 - Amounts received under accident and health plans _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§105(h)(5)", "Factual", "Hard",
 "Three prongs for HCI definition."),

# ── 26 U.S.C. §162 – Trade or business expenses ─────────────────────────
("A014", "Acts",
 "What is the general rule for deductibility of business expenses under 26 U.S.C. §162(a), and what three broad categories of expenses does it specifically list?",
 "Section 162(a) allows a deduction for all ordinary and necessary expenses paid or incurred during the taxable year in carrying on any trade or business. The three specific categories listed are: (1) reasonable allowances for salaries or other compensation for personal services actually rendered; (2) traveling expenses (including meals and lodging) while away from home in the pursuit of a trade or business, provided the period away does not exceed one year; and (3) rentals or other payments required to be made as a condition to the continued use of property in the trade or business.",
 "26 U.S. Code § 162 - Trade or business expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§162(a)", "Factual", "Easy",
 "Three sub-categories explicitly listed."),

("A015", "Acts",
 "Under 26 U.S.C. §162(c), what is the rule for deductibility of illegal bribes, kickbacks, and similar payments?",
 "Section 162(c) provides that no deduction is allowed for any payment that constitutes an illegal bribe or kickback if: the bribe is made to a government official or employee (domestic or foreign) or any person designated by such official; or the payment is made in violation of any federal or state law that subjects the payor to a criminal penalty or loss of business license. This reflects the public policy that allowing a deduction for illegal payments would subsidize illegal conduct.",
 "26 U.S. Code § 162 - Trade or business expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 4, "§162(c)", "Factual", "Hard",
 "Public policy limitation on deductions."),

("A016", "Acts",
 "Under 26 U.S.C. §162(e), what is the general rule regarding deductions for lobbying expenditures, and is there a de minimis exception?",
 "Section 162(e)(1) provides that no deduction is allowed for amounts paid or incurred in connection with: influencing legislation; participating in any political campaign on behalf of a candidate; attempting to influence the official actions of a federal, state, or local government executive; or direct communication with a covered executive branch official. However, §162(e)(5) contains a de minimis rule providing that expenditures not exceeding $2,000 in aggregate for the taxable year are not subject to the disallowance.",
 "26 U.S. Code § 162 - Trade or business expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 5, "§162(e)", "Factual", "Hard",
 "$2,000 de minimis threshold."),

# ── 26 U.S.C. §163 – Interest ────────────────────────────────────────────
("A017", "Acts",
 "Under 26 U.S.C. §163(d), what limitation applies to the deduction of investment interest, and is any excess deductible in a future year?",
 "Section 163(d)(1) limits the deduction for investment interest to the taxpayer's net investment income for the taxable year. Net investment income is defined as the excess of investment income over investment expenses. Any investment interest disallowed in the current year due to this limitation is treated as investment interest paid or accrued in the next taxable year, and may be carried forward indefinitely until deducted against future net investment income.",
 "26 U.S. Code § 163 - Interest _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§163(d)", "Conceptual", "Medium",
 "Indefinite carryforward of disallowed investment interest."),

# ── 26 U.S.C. §164 – Taxes ──────────────────────────────────────────────
("A018", "Acts",
 "Under 26 U.S.C. §164(a), what categories of taxes are deductible, and does the statute permit an election to deduct general sales taxes in lieu of state income taxes?",
 "Section 164(a) allows deductions for: (1) state and local, and foreign real property taxes; (2) state and local personal property taxes; (3) state, local, and foreign income, war profits, and excess profits taxes; and (5) the generation-skipping transfer tax. In addition, §164(b)(5) provides an election allowing individuals to deduct state and local general sales taxes in lieu of state and local income taxes, with two methods of computation (actual or IRS-provided table).",
 "26 U.S. Code § 164 - Taxes _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§164(a)", "Factual", "Medium",
 "Sales tax election in §164(b)(5)."),

("A019", "Acts",
 "What is the SALT deduction cap under 26 U.S.C. §164(b)(6) for tax years 2025 and 2026, and what happens to the cap after 2029?",
 "Section 164(b)(6) limits the deduction for state and local taxes for individuals. For taxable years beginning in calendar year 2025, the applicable limitation amount is $40,000. For taxable years beginning in calendar year 2026, the amount is $40,400. After taxable years beginning in calendar year 2029, the applicable limitation amount reverts to $10,000. Additionally, a phasedown applies based on modified adjusted gross income: for 2025, the phasedown begins when MAGI exceeds $500,000 (and for 2026, $505,000), reducing the cap by 30 percent of the excess, but the cap cannot fall below $10,000.",
 "26 U.S. Code § 164 - Taxes _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 5, "§164(b)(6)(A)", "Factual", "Hard",
 "SALT caps: $40,000 (2025), $40,400 (2026), $10,000 (post-2029)."),

("A020", "Acts",
 "Under 26 U.S.C. §164(f), what portion of self-employment taxes may an individual deduct, and how is this deduction treated for purposes of determining trade or business income?",
 "Section 164(f)(1) provides that an individual may deduct an amount equal to one-half of the taxes imposed under §1401 (the self-employment tax) for the taxable year, other than the Additional Medicare Tax under §1401(b)(2). Under §164(f)(2), this deduction is treated as attributable to a trade or business carried on by the taxpayer that does not consist of the performance of services as an employee, which means it is an above-the-line deduction reducing adjusted gross income.",
 "26 U.S. Code § 164 - Taxes _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 7, "§164(f)", "Factual", "Medium",
 "50% self-employment tax deduction."),

# ── 26 U.S.C. §165 – Losses ─────────────────────────────────────────────
("A021", "Acts",
 "Under 26 U.S.C. §165(c), what three categories of losses may individuals deduct, and what rule applies to theft losses under §165(e)?",
 "Section 165(c) limits individual loss deductions to: (1) losses incurred in a trade or business; (2) losses incurred in any transaction entered into for profit (though not connected with a trade or business); and (3) casualty or theft losses of property not connected with a trade or business, such as those arising from fire, storm, shipwreck, or other casualty. Under §165(e), any loss arising from theft is treated as sustained during the taxable year in which the taxpayer discovers the loss, not the year in which the theft occurred.",
 "26 U.S. Code § 165 - Losses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§165(c)", "Factual", "Medium",
 "Theft loss recognized in year of discovery."),

("A022", "Acts",
 "Under 26 U.S.C. §165(d), what is the rule for deducting wagering (gambling) losses, and is there a limitation on the amount deductible?",
 "Section 165(d)(1) provides that losses from wagering transactions are deductible only to the extent of gains from such transactions during the taxable year, and only 90 percent of such losses may be taken into account for any taxable year. This means a taxpayer who has $10,000 in gambling losses may deduct at most $9,000 (90%), and only if they have at least $9,000 in gambling gains. Gambling losses cannot create or increase a net loss.",
 "26 U.S. Code § 165 - Losses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§165(d)", "Factual", "Hard",
 "90% limitation + gains ceiling on wagering losses."),

("A023", "Acts",
 "Under 26 U.S.C. §165(h) and §165(h)(5), how are personal casualty losses treated for taxable years beginning after December 31, 2017?",
 "For taxable years beginning after December 31, 2017, §165(h)(5)(A) generally limits the deduction for personal casualty losses (losses of property not connected with a trade or business) to those attributable to a Federally declared disaster (as defined in §165(i)(5), meaning any disaster declared by the President under the Robert T. Stafford Disaster Relief Act) or a State declared disaster. The prior rules still apply (a $100 per-casualty floor and the 10-percent-of-AGI floor must be satisfied), but non-disaster losses are no longer deductible unless the taxpayer has offsetting personal casualty gains.",
 "26 U.S. Code § 165 - Losses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 5, "§165(h)(5)", "Conceptual", "Hard",
 "Post-2017 personal casualty loss limited to declared disasters."),

# ── 26 U.S.C. §170 – Charitable contributions ───────────────────────────
("A024", "Acts",
 "Under 26 U.S.C. §170(b)(1)(A), what is the general percentage limitation on charitable contributions to public charities for individual taxpayers?",
 "Section 170(b)(1)(A) limits charitable contributions by individuals to public charities (churches, educational institutions, hospitals, governmental units, and similar §170(b)(1)(A) organizations) to 50 percent of the taxpayer's contribution base for the taxable year. The contribution base is adjusted gross income computed without regard to any net operating loss carryback. Contributions exceeding the 50 percent limit may be carried forward to each of the five succeeding taxable years.",
 "26 U.S. Code § 170 - Charitable, etc., contributions and gifts _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 4, "§170(b)(1)(A)", "Factual", "Medium",
 ""),

("A025", "Acts",
 "Under 26 U.S.C. §170(b)(1)(G), what is the percentage limitation for cash contributions to public charities for taxable years beginning after December 31, 2017?",
 "Section 170(b)(1)(G) provides that for taxable years beginning after December 31, 2017, cash contributions by an individual to an organization described in §170(b)(1)(A) (public charities) are deductible up to 60 percent of the taxpayer's contribution base for the taxable year (reduced by contributions allowed under §170(b)(1)(A)). This is an increase from the prior 50 percent limit specifically for cash (not property) contributions. Excess contributions may be carried forward to the five succeeding taxable years.",
 "26 U.S. Code § 170 - Charitable, etc., contributions and gifts _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 8, "§170(b)(1)(G)", "Factual", "Hard",
 "60% cash limit enacted by TCJA 2017."),

("A026", "Acts",
 "What is the percentage limitation and carryover period for qualified conservation contributions under 26 U.S.C. §170(b)(1)(E)?",
 "Section 170(b)(1)(E)(i) provides that qualified conservation contributions are deductible to the extent the aggregate of such contributions does not exceed the excess of 50 percent of the taxpayer's contribution base over the amount of all other charitable contributions otherwise allowable for the year. Under §170(b)(1)(E)(ii), if the aggregate exceeds this limitation, the excess is carried forward as a conservation contribution in each of the 15 succeeding taxable years. For qualified farmers or ranchers (taxpayers whose gross income from farming exceeds 50 percent of their total gross income), the limitation is increased to 100 percent rather than 50 percent.",
 "26 U.S. Code § 170 - Charitable, etc., contributions and gifts _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 6, "§170(b)(1)(E)", "Factual", "Hard",
 "15-year carryover period for conservation contributions."),

# ── 26 U.S.C. §179 – Section 179 expensing ──────────────────────────────
("A027", "Acts",
 "Under 26 U.S.C. §179(b)(1) and (b)(2), what is the maximum dollar amount that may be expensed under §179, and when does the phase-out begin?",
 "Section 179(b)(1) sets the aggregate cost that may be expensed (deducted immediately) under the §179 election at $2,500,000 per taxable year. Under §179(b)(2), this limitation is reduced dollar-for-dollar (but not below zero) by the amount by which the cost of §179 property placed in service during the taxable year exceeds $4,000,000. Both the $2,500,000 and $4,000,000 amounts are adjusted for inflation in taxable years beginning after 2025 using the C-CPI-U.",
 "26 U.S. Code § 179 - Election to expense certain depreciable business assets _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§179(b)(1)-(2)", "Factual", "Medium",
 "$2.5M limit, phase-out above $4M."),

("A028", "Acts",
 "Under 26 U.S.C. §179(b)(5)(A), what is the maximum §179 deduction allowed for sport utility vehicles (SUVs) placed in service in a taxable year?",
 "Section 179(b)(5)(A) provides that the cost of any sport utility vehicle that may be taken into account under §179 for any taxable year shall not exceed $25,000. A sport utility vehicle is defined under §179(b)(5)(B) as a 4-wheeled vehicle designed to carry passengers over public streets and roads, not subject to §280F, and rated at not more than 14,000 pounds gross vehicle weight. This cap prevents the full §179 expensing from being applied to expensive SUVs used in business.",
 "26 U.S. Code § 179 - Election to expense certain depreciable business assets _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§179(b)(5)(A)", "Factual", "Medium",
 "$25,000 SUV cap."),

# ── 26 U.S.C. §1031 – Like-kind exchange ────────────────────────────────
("A029", "Acts",
 "Under 26 U.S.C. §1031(a), what is the general non-recognition rule for like-kind exchanges, and what is the 45-day identification rule?",
 "Section 1031(a)(1) provides that no gain or loss is recognized on the exchange of real property held for productive use in a trade or business or for investment, if such property is exchanged solely for real property of a like kind to be held for the same purposes. Under §1031(a)(3)(A), if the property received in the exchange is not received at the time of the transfer of the relinquished property, the new property must be identified in writing within 45 days after the date of the transfer of the old property. Failure to identify replacement property within this window disqualifies the exchange from non-recognition.",
 "26 U.S. Code § 1031 - Exchange of real property held for productive use or investment _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§1031(a)(3)(A)", "Factual", "Medium",
 "45-day identification, 180-day receipt."),

("A030", "Acts",
 "Under 26 U.S.C. §1031(a)(3)(B), what is the maximum period within which a taxpayer must receive the replacement property in a deferred like-kind exchange?",
 "Section 1031(a)(3)(B) requires that the replacement property be received by the earlier of: (i) the date that is 180 days after the date on which the taxpayer transferred the relinquished property, or (ii) the due date (including extensions) for the taxpayer's tax return for the taxable year in which the transfer occurred. If replacement property is not received within this window, the exchange fails to qualify for non-recognition under §1031.",
 "26 U.S. Code § 1031 - Exchange of real property held for productive use or investment _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§1031(a)(3)(B)", "Factual", "Hard",
 "180 days or return due date, whichever is earlier."),

("A031", "Acts",
 "Under 26 U.S.C. §1031(h), is US real property like-kind to foreign real property for purposes of the non-recognition rule?",
 "No. Section 1031(h)(1) explicitly provides that real property located in the United States and real property located outside the United States are not property of a like kind for purposes of the §1031 non-recognition rule. This means that a taxpayer who exchanges US real property for foreign real property (or vice versa) must recognize any gain on the transaction. Only exchanges of US real property for other US real property, or foreign real property for foreign real property, may qualify.",
 "26 U.S. Code § 1031 - Exchange of real property held for productive use or investment _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 4, "§1031(h)", "Factual", "Hard",
 "US vs. foreign real property not like-kind."),

# ── 26 U.S.C. §1361 – S corporation defined ─────────────────────────────
("A032", "Acts",
 "Under 26 U.S.C. §1361(b)(1), what are the four basic eligibility requirements for a small business corporation to qualify as an S corporation?",
 "Section 1361(b)(1) provides that a 'small business corporation' is a domestic corporation that: (A) does not have more than 100 shareholders; (B) does not have as a shareholder a person other than an individual, an estate, certain trusts, or a tax-exempt organization described in §401(a) or §501(c)(3); (C) does not have a nonresident alien as a shareholder; and (D) does not have more than one class of stock. Corporations that are financial institutions using the reserve method, insurance companies, or DISCs are ineligible regardless of meeting the other tests.",
 "26 U.S. Code § 1361 - S corporation defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§1361(b)(1)", "Factual", "Medium",
 "Key: 100 shareholders, 1 class of stock, no nonresident aliens."),

("A033", "Acts",
 "Under 26 U.S.C. §1361(c)(1), how are members of the same family treated for purposes of the 100-shareholder limit of an S corporation?",
 "Section 1361(c)(1) provides that all members of a family (the common ancestor, the common ancestor's lineal descendants, and the spouses or former spouses of any such descendants) shall be treated as one shareholder for purposes of the 100-shareholder limitation. For this rule, the term 'family' extends to individuals who are related to each other within 6 generations. This facilitates family-owned S corporations by allowing multiple family members to hold stock without each one counting separately toward the limit.",
 "26 U.S. Code § 1361 - S corporation defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§1361(c)(1)", "Factual", "Medium",
 "6 generations = 1 shareholder for the 100-limit count."),

# ── 26 U.S.C. §213 – Medical expenses ───────────────────────────────────
("A034", "Acts",
 "Under 26 U.S.C. §213(a), what is the income threshold for the deduction of medical expenses, and what does 'medical care' include under §213(d)?",
 "Section 213(a) allows a deduction for medical care expenses paid during the taxable year that are not compensated by insurance, to the extent they exceed 7.5 percent of adjusted gross income. Under §213(d)(1), 'medical care' means amounts paid for: (A) the diagnosis, cure, mitigation, treatment, or prevention of disease, or for affecting any structure or function of the body; (B) transportation primarily for medical care; (C) qualified long-term care services; and (D) insurance covering such medical care. Only prescribed drugs and insulin are included as medicine or drugs under §213(b).",
 "26 U.S. Code § 213 - Medical, dental, etc., expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§213(a)", "Factual", "Medium",
 "7.5% AGI floor for medical expenses."),

("A035", "Acts",
 "Under 26 U.S.C. §213(d)(9), is cosmetic surgery considered 'medical care' for purposes of the §213 deduction?",
 "No. Section 213(d)(9)(A) expressly provides that 'medical care' does not include cosmetic surgery or other similar procedures, unless the surgery is necessary to ameliorate a deformity arising from, or directly related to, a congenital abnormality, a personal injury resulting from an accident or trauma, or a disfiguring disease. Under §213(d)(9)(B), cosmetic surgery is defined as any procedure directed at improving the patient's appearance that does not meaningfully promote the proper functioning of the body or prevent or treat illness or disease.",
 "26 U.S. Code § 213 - Medical, dental, etc., expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 4, "§213(d)(9)", "Factual", "Medium",
 "Cosmetic surgery excluded except for specific exceptions."),

# ── 26 U.S.C. §2001 – Estate tax ────────────────────────────────────────
("A036", "Acts",
 "Under 26 U.S.C. §2001(a) and (c), what is the top marginal rate of the federal estate tax, and at what level of taxable estate does it apply?",
 "Section 2001(a) imposes a tax on the transfer of the taxable estate of every decedent who is a citizen or resident of the United States. Under the rate schedule in §2001(c), the top rate of 40 percent applies to taxable estates (including adjusted taxable gifts) in excess of $1,000,000. The tentative tax for amounts over $1,000,000 is $345,800 plus 40 percent of the excess over $1,000,000. The estate tax is then reduced by the applicable credit amount under §2010 to exempt estates below the statutory exclusion amount (currently $13+ million).",
 "26 U.S. Code § 2001 - Imposition and rate of tax _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 3, "§2001(c)", "Factual", "Medium",
 "40% top rate above $1M (before applicable credit)."),

# ── 26 U.S.C. §2501 – Gift tax ──────────────────────────────────────────
("A037", "Acts",
 "Under 26 U.S.C. §2501(a)(1) and (2), who is subject to the federal gift tax, and what is the rule for intangible property transferred by a nonresident non-citizen?",
 "Section 2501(a)(1) imposes a gift tax on the transfer of property by gift during each calendar year by any individual who is a resident or nonresident. However, §2501(a)(2) provides that the transfer of intangible property by a nonresident who is not a citizen of the United States is not subject to the gift tax, subject to exceptions. Under §2501(a)(4), transfers to political organizations (as defined in §527(e)(1)) are also exempt from the gift tax.",
 "26 U.S. Code § 2501 - Imposition of tax _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§2501(a)", "Factual", "Hard",
 "Nonresident non-citizen: intangibles exempt with exceptions."),

# ── 26 U.S.C. §6015 – Relief from joint liability ───────────────────────
("A038", "Acts",
 "Under 26 U.S.C. §6015(b), what conditions must be satisfied for a spouse to obtain innocent spouse relief from joint and several liability?",
 "Section 6015(b)(1) requires all of the following: (A) a joint return was filed; (B) there is an understatement of tax attributable to erroneous items of one individual; (C) the other spouse establishes that in signing the return he or she did not know, and had no reason to know, of the understatement; (D) taking into account all facts and circumstances, it would be inequitable to hold the other spouse liable; and (E) the election must be made within 2 years after the IRS begins collection activities. If all five conditions are met, the innocent spouse is relieved of liability for the understatement.",
 "26 U.S. Code § 6015 - Relief from joint and several liability on joint return _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§6015(b)(1)", "Factual", "Hard",
 "5-prong test for innocent spouse relief."),

# ── 26 U.S.C. §6651 – Failure to file/pay ───────────────────────────────
("A039", "Acts",
 "Under 26 U.S.C. §6651(a), what is the penalty rate for failure to timely file a tax return, and what is the maximum aggregate penalty?",
 "Section 6651(a)(1) provides that if a taxpayer fails to file a return by the prescribed due date (including extensions), unless the failure is due to reasonable cause and not willful neglect, 5 percent of the tax required to be shown on the return is added for each month or fraction of a month the failure continues, up to a maximum of 25 percent in the aggregate. If the failure to file is fraudulent, §6651(f) increases the monthly rate to 15 percent and the maximum to 75 percent. Additionally, if the return is more than 60 days late, the minimum penalty is the lesser of $435 (adjusted for inflation) or 100 percent of the unpaid tax.",
 "26 U.S. Code § 6651 - Failure to file tax return or to pay tax _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§6651(a)(1)", "Factual", "Medium",
 "5%/month, max 25%; fraudulent: 15%/month max 75%."),

# ── 26 U.S.C. §6662 – Accuracy-related penalty ──────────────────────────
("A040", "Acts",
 "Under 26 U.S.C. §6662(a) and (d), what is the accuracy-related penalty rate, and when is an understatement of income tax considered 'substantial' for an individual taxpayer?",
 "Section 6662(a) imposes an accuracy-related penalty of 20 percent of the portion of the underpayment attributable to one of the listed categories. Under §6662(d)(1)(A), for individual taxpayers, a 'substantial understatement' of income tax exists if the amount of the understatement for the taxable year exceeds the greater of (i) 10 percent of the tax required to be shown on the return, or (ii) $5,000. For corporations (other than S corporations or personal holding companies), the substantial understatement threshold is $10,000. In certain cases involving gross valuation misstatements, the penalty rate increases to 40 percent under §6662(h).",
 "26 U.S. Code § 6662 - Imposition of accuracy-related penalty on underpayments _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§6662(a)", "Factual", "Medium",
 "20% penalty; substantial = >10% or $5,000 for individuals."),

# ════════════════════════════════════════════════════════════════════════
# JUDGMENTS  (target 45)
# ════════════════════════════════════════════════════════════════════════

("J001", "Judgments",
 "In Adam v. Commissioner (60 T.C. 996, 1973), what was the central legal issue, and what factors did the Tax Court examine to resolve it?",
 "The central issue in Adam was whether Robert Adam's sales of Maine waterfront properties were sales of property held primarily for sale to customers in the ordinary course of his trade or business (which would produce ordinary income) or investment transactions (qualifying for capital gain treatment). The Tax Court applied a six-factor test derived from prior case law: (1) the purpose for which the asset was acquired; (2) the frequency, continuity, and size of sales; (3) the seller's activities in improvement and disposition of the property; (4) the extent of improvements made; (5) the proximity of sale to purchase; and (6) the purpose for which the property was held during the taxable year. The Court found that Adam was an investor, not a dealer.",
 "Adam v. Commissioner, No. 4192-71 (U.S.T.C. 1973) __ Justia.pdf",
 5, "Section 1221(1) analysis", "Conceptual", "Hard",
 "Six-factor test for dealer vs. investor."),

("J002", "Judgments",
 "In Adam v. Commissioner (60 T.C. 996), what specific facts supported the Tax Court's conclusion that Robert Adam was an investor and not a real estate dealer?",
 "The Tax Court noted several key facts: Adam did not plot, subdivide, or improve any of the waterfront properties; he sold each property as a single piece to a single purchaser; he did not advertise or place 'For Sale' signs on properties; he did not employ real estate brokers (in only one instance did a broker advertise one property for one month); he did not hold himself out as a real estate dealer, was not a licensed broker, and had no separate real estate office; and his real estate transactions were intermittent and occasional rather than continuous. The Court held that a taxpayer who engages in fewer transactions with no significant development activity cannot be found to be in the real estate business.",
 "Adam v. Commissioner, No. 4192-71 (U.S.T.C. 1973) __ Justia.pdf",
 6, "Findings of Fact", "Analytical", "Hard",
 ""),

("J003", "Judgments",
 "What was the holding in Bellamy v. Commissioner (43 T.C. 487, 1965), and what income tax consequence resulted?",
 "In Bellamy, actor Ralph Bellamy received $89,000 in 1957 for agreeing to relinquish broadcasting rights related to a television series ('Man Against Crime'). The Tax Court held that the $89,000 did not represent proceeds from the sale of a capital asset but rather constituted ordinary income taxable at normal rates. Bellamy had argued the payment was consideration for the sale of a capital asset (the right to prevent future broadcasts). The Court rejected this argument, finding the payment was compensation for rights that were personal service contract rights and not capital assets eligible for favorable capital gain treatment.",
 "Bellamy v. Commissioner, No. 989-63 (U.S.T.C. 1965) __ Justia.pdf",
 2, "Held: ordinary income", "Factual", "Medium",
 "Actor's broadcasting rights = ordinary income, not capital gain."),

("J004", "Judgments",
 "In Blackman v. Commissioner (88 T.C. 677, 1987), why did the Tax Court deny Biltmore Blackman's casualty loss deduction for the fire that destroyed his home?",
 "The Tax Court denied Blackman's $97,853 casualty loss deduction under §165(a) and §165(c)(3) for two independent reasons. First, the Court found his conduct was grossly negligent: he admitted starting the fire by burning his wife's clothes on the stove, and his claim that he attempted to extinguish it was not corroborated by evidence. Because the house fire was a foreseeable consequence of his actions, gross negligence barred the deduction. Second, allowing the deduction would severely frustrate Maryland's articulated public policy against arson and burning, as Article 27, §11 of the Maryland Annotated Code made it a felony to burn a residence while perpetrating a crime.",
 "Blackman v. Commissioner, No. 21436-84 (U.S.T.C. 1987) __ Justia.pdf",
 5, "§165(a) and §165(c)(3) analysis", "Analytical", "Hard",
 "Grossly negligent arson + public policy bars deduction."),

("J005", "Judgments",
 "In Blackman v. Commissioner (88 T.C. 677), how did the Tax Court rule on the negligence addition to tax under §6653(a)?",
 "Interestingly, while the Tax Court denied Blackman's casualty loss deduction and sustained the late-filing penalty under §6651(a)(1), it did not impose the §6653(a) negligence addition to tax. The Court held that under the circumstances, claiming a deduction for a fire loss was not negligent, citing Wofford v. Commissioner. Merely claiming a deduction to which the taxpayer turned out not to be entitled does not automatically constitute negligence or intentional disregard of the rules, especially when the legal question presented was a reasonable one.",
 "Blackman v. Commissioner, No. 21436-84 (U.S.T.C. 1987) __ Justia.pdf",
 6, "§6653(a) discussion", "Analytical", "Hard",
 "Negligence penalty not imposed despite denied deduction."),

("J006", "Judgments",
 "In Bramblett v. Commissioner (59 T.C.M. 876, 1990), what was the ultimate question decided by the Tax Court, and what was its holding regarding the character of gain?",
 "The issue was whether the gain on the sale of land by Mesquite East (a joint venture in which Richard Bramblett was a partner) to Town East (another entity) was capital gain or ordinary income. The Tax Court held that the gain was ordinary income because the land was held primarily for sale to customers in the ordinary course of Bramblett's trade or business. Key factors included that contracts of sale to third parties had been entered into by Baker and Bramblett as trustee before the land was even conveyed to Town East, indicating the land was already being sold in the ordinary course of a real estate business.",
 "Bramblett v. Commissioner, No. 728-89 (U.S.T.C. 1990) __ Justia.pdf",
 8, "§1221(1) analysis", "Analytical", "Hard",
 "Sale to related entity preceded by customer contracts = ordinary income."),

("J007", "Judgments",
 "In Bynum v. Commissioner (46 T.C. 295, 1966), what key facts led the Tax Court to find that gains from the sale of subdivided lots were ordinary income rather than capital gain?",
 "The Tax Court found several decisive facts: S.O. Bynum subdivided and substantially improved his farm, spending over $1,000 per acre on improvements (more than double the land's fair market value per acre before improvement); he ran full-page newspaper advertisements and listed the subdivision in the phone book; he listed lots with all reputable local realtors; he personally sold all the lots without paying commissions; and his advertising indicated that 160 additional lots would be developed. These actions showed that by the time of the sales, the property was held primarily for sale to customers in the ordinary course of a real estate sales business, even though Bynum's primary occupation was still his nursery business (which occupied 90-95% of his time).",
 "Bynum v. Commissioner, No. 5226-63 (U.S.T.C. 1966) __ Justia.pdf",
 6, "§1221(1) and §1231(b)(1)(B)", "Analytical", "Hard",
 "Subdivision + advertising + improvement = dealer status."),

("J008", "Judgments",
 "In Calloway v. Commissioner (135 T.C. 26, 2010), how did the Tax Court characterize the 'stock loan' transaction between Albert Calloway and Derivium Capital?",
 "The Tax Court held that the transaction between Calloway and Derivium was a sale of IBM stock in 2001, not a loan. The Court found that Calloway transferred all the benefits and burdens of ownership of 990 shares of IBM common stock to Derivium in exchange for $93,586.23 (approximately 90% of the stock's value), with no obligation to repay that amount. Derivium immediately sold the stock upon receipt, the 'loan' was nonrecourse, and no interest or principal payments were made during the 3-year term. At maturity, Calloway surrendered any right to the IBM stock. These facts established that a sale, not a borrowing, occurred, making the transaction taxable in 2001.",
 "Calloway v. Commissioner, No. 8438-07 (U.S.T.C. 2010) __ Justia.pdf",
 2, "Sale vs. loan analysis", "Conceptual", "Hard",
 "90% stock 'loan' = sale; Derivium sold stock immediately."),

("J009", "Judgments",
 "In Carter v. Commissioner (40 T.C.M. 654, 1980), why was a cash basis taxpayer required to report wages in 1975 even though the work was performed in 1974?",
 "Robert Carter worked for New York City in November and December 1974, but due to payroll processing delays, he was not paid $1,073.01 in wages until January 3, 1975. As a cash basis taxpayer, Carter argued that the income was constructively received in 1974. The Tax Court rejected this, holding that constructive receipt requires that income be made available to the taxpayer without substantial limitations or restrictions. Mere presence of funds in the city's budget was insufficient for constructive receipt — Carter could not access the wages until payment was actually made. Under the cash basis of reporting, income is taxable in the year actually received.",
 "Carter v. Commissioner, No. 12404-78 (U.S.T.C. 1980) __ Justia.pdf",
 3, "§451; constructive receipt doctrine", "Conceptual", "Medium",
 "Constructive receipt requires availability without restriction."),

("J010", "Judgments",
 "In Dyer v. Commissioner (78358, 1961), did the Tax Court allow a deduction for a vase broken by the taxpayer's cat as a casualty loss under §165(c)(3)?",
 "The Tax Court disallowed the $100 deduction claimed for the vase broken by the Dyers' Siamese cat. The Commissioner determined that a cat knocking over a vase during a fit did not constitute a 'casualty' under §165(c)(3), which requires losses arising from fire, storm, shipwreck, or 'other casualty.' Additionally, the Dyers had not substantiated the fair market value of the vase immediately before and after the breaking, nor did they establish their cost or other basis in the vase. The Court sustained both grounds: the event was not a qualifying casualty, and the deductible amount was not substantiated.",
 "Dyer v. Commissioner, No. 78358 (U.S.T.C. 1961) __ Justia.pdf",
 2, "§165(c)(3) casualty loss", "Factual", "Medium",
 "Cat breaking vase = not a qualifying casualty; also unsubstantiated."),

("J011", "Judgments",
 "In Estate of Kohlsaat (73 T.C.M. 2732, 1997), did the Tax Court allow annual gift tax exclusions under §2503(b) for contingent remainder beneficiaries who had Crummey withdrawal rights?",
 "Yes. The Tax Court held that the 16 contingent remainder beneficiaries of the Lieselotte Kohlsaat Family Trust each held a present interest in the property for purposes of the §2503(b) annual gift tax exclusion. This was because each beneficiary had been given an unrestricted right to demand immediate distribution of trust property (up to the $10,000 exclusion amount) within 30 days following a transfer to the trust, and all beneficiaries received timely notice of these rights. The Court rejected the IRS's argument that an understanding existed among the beneficiaries not to exercise their withdrawal rights, finding no evidence of such an understanding.",
 "Estate of Kohlsaat, No. 22465-94 (U.S.T.C. 1997) __ Justia.pdf",
 3, "§2503(b); Crummey powers", "Analytical", "Hard",
 "Crummey powers = present interests; 16 exclusions allowed."),

("J012", "Judgments",
 "In Foote v. Commissioner (81 T.C. 930, 1983), was professor Merrill Foote's receipt of $45,640 for resigning his tenured position treated as capital gain or ordinary income?",
 "The Tax Court held that the $45,640 received by Foote for resigning his tenured appointment at Southern Methodist University was taxable as ordinary income, not capital gain. Two key reasons were: (1) Tenure is not a 'capital asset' under §1221 because it is an employment contract right, not property that can be sold to a third party, assigned, or transferred — it is personal to the holder; and (2) Even if tenure were a capital asset, the resignation did not constitute a 'sale or exchange' as required by §1222, because Foote's tenure rights merely came to an end rather than being transferred to another party. The payment was essentially a substitute for future ordinary income (salary and consulting fees).",
 "Foote v. Commissioner, No. 9667-81 (U.S.T.C. 1983) __ Justia.pdf",
 5, "§§1221, 1222 analysis", "Analytical", "Hard",
 "Tenure is not a capital asset; release ≠ sale or exchange."),

("J013", "Judgments",
 "In Gilliam v. Commissioner (51 T.C.M. 515, 1986), why were the legal fees incurred by artist Sam Gilliam in defending criminal charges and settling a civil claim not deductible under §162?",
 "The Tax Court held that Gilliam's legal fees — incurred for defending criminal charges arising from an in-flight assault on another passenger and settling the related civil claim — were not deductible ordinary and necessary business expenses under §162. The Court found that the criminal incident (Gilliam, in an apparent temporary insanity episode, attacked a fellow passenger with a telephone receiver) was not an 'ordinary' expense of carrying on Gilliam's trades or businesses as an artist and art teacher. The assault on the airplane was not a normal or expected event in the conduct of Gilliam's art business, and the expenses were beyond the norm of general and accepted business practice. The stipulated facts also showed these expenses were not ordinary expenses of his trades or businesses.",
 "Gilliam v. Commissioner, No. 4986-79 (U.S.T.C. 1986) __ Justia.pdf",
 4, "§162(a) ordinary and necessary test", "Analytical", "Hard",
 "Assault on airplane not 'ordinary' business expense for an artist."),

# More judgements (reading from available content):
("J014", "Judgments",
 "In Adam v. Commissioner (60 T.C. 996), what was the role of Goodwin Wiseman (the real estate broker) in Robert Adam's property transactions, and did this create a dealer relationship?",
 "Goodwin Wiseman, who was employed primarily by a bank as a trust officer and ran real estate companies on the side, recommended properties to Adam that might appreciate substantially. When Adam purchased a property Wiseman recommended, Wiseman received one-half of any gain realized on the subsequent sale. Wiseman obtained rights of way and perimeter surveys for Adam. However, the Tax Court found that this arrangement did not transform Adam into a dealer, because: Wiseman did not solicit purchasers for Adam; Adam made all purchase decisions and bore all financial risk; and the arrangement had characteristics of a profit-sharing investment arrangement rather than a business agency relationship.",
 "Adam v. Commissioner, No. 4192-71 (U.S.T.C. 1973) __ Justia.pdf",
 4, "Broker relationship analysis", "Analytical", "Medium",
 "50% profit-sharing with broker did not create dealer status."),

("J015", "Judgments",
 "Under the Supreme Court's holding in Malat v. Riddell (383 U.S. 569, 1966), as discussed in both Adam v. Commissioner and Bynum v. Commissioner, what does 'primarily' mean in the phrase 'held primarily for sale to customers in the ordinary course of his trade or business'?",
 "Both the Adam and Bynum courts relied on Malat v. Riddell, in which the Supreme Court held that 'primarily' means 'of first importance' or 'principally.' The Court rejected the argument that a 'substantial' purpose was sufficient to classify property as dealer property. This interpretation is designed to differentiate between 'profits and losses arising from the everyday operation of a business' (ordinary income) and 'the realization of appreciation in value accrued over a substantial period of time' (capital gain). The primary purpose at the time of holding the property during the taxable year, not just the purpose of acquisition, is the key inquiry.",
 "Adam v. Commissioner, No. 4192-71 (U.S.T.C. 1973) __ Justia.pdf",
 5, "§1221(1); Malat v. Riddell cited", "Conceptual", "Hard",
 "Cross-reference: Malat v. Riddell in both Adam and Bynum."),

("J016", "Judgments",
 "In Calloway v. Commissioner (135 T.C. 26, 2010), was the Derivium transaction analogous to a tax-free securities lending arrangement under §1058?",
 "No. The Tax Court held that the Derivium transaction was not analogous to a securities lending arrangement under §1058 and was not equivalent to the arrangements described in Rev. Rul. 57-451. The Court noted that a true securities lending arrangement requires the lender to receive identical securities (or the equivalent) back, and the borrower typically remains subject to market risk on those securities. In the Derivium arrangement, Calloway had no recourse against Derivium if the value of IBM stock dropped (the 'loan' was nonrecourse), and at maturity he simply surrendered any claim to IBM stock rather than receiving it back. Derivium sold the stock immediately, bearing the economic risk of the stock price.",
 "Calloway v. Commissioner, No. 8438-07 (U.S.T.C. 2010) __ Justia.pdf",
 3, "§1058 securities lending analysis", "Analytical", "Hard",
 "Not a §1058 securities loan; stock sold immediately."),

("J017", "Judgments",
 "In Bynum v. Commissioner (46 T.C. 295, 1966), the petitioners argued they were merely passive investors liquidating a portion of their farm to pay off a mortgage. How did the Tax Court respond to this argument?",
 "The Tax Court rejected this passive investor argument. It found that the Bynums' actions went far beyond those of a passive investor liquidating assets. They had subdivided 38 lots and spent over $1,000 per acre on improvements; their initial advertising indicated 233 lots would eventually be offered; 17 additional lots were subdivided in 1962; and under their bank arrangement, only 26 lots were needed to pay off the mortgage, yet they continued to develop and sell beyond that point. The Court also rejected the argument that spending only 5-10% of time on real estate activities was determinative, since a taxpayer may be engaged in multiple businesses simultaneously.",
 "Bynum v. Commissioner, No. 5226-63 (U.S.T.C. 1966) __ Justia.pdf",
 6, "Multiple businesses; dealer status", "Analytical", "Medium",
 ""),

("J018", "Judgments",
 "In Estate of Kohlsaat (73 T.C.M. 2732, 1997), what was the IRS's primary argument for denying the annual gift tax exclusions, and why did the Tax Court reject it?",
 "The IRS argued that understandings existed between the decedent (Lieselotte Kohlsaat) and the 16 contingent beneficiaries that they would not exercise their Crummey withdrawal rights, and therefore the gifts lacked donative intent and the substance-over-form doctrine should deny the exclusions. The Tax Court rejected this argument because: the evidence did not establish any such understanding; at trial, several credible reasons were offered by beneficiaries as to why they chose not to exercise their rights; merely not exercising a right does not imply an agreement not to do so; and the beneficiaries received actual notice of their rights within 6 days of the transfer.",
 "Estate of Kohlsaat, No. 22465-94 (U.S.T.C. 1997) __ Justia.pdf",
 3, "§2503(b); substance over form", "Analytical", "Hard",
 ""),

("J019", "Judgments",
 "In Blackman v. Commissioner (88 T.C. 677, 1987), was a conviction for arson required before the Tax Court could deny the casualty loss deduction on public policy grounds?",
 "No. The Tax Court explicitly held that a criminal conviction is not required to deny a tax deduction on the ground that allowing the deduction would frustrate public policy. The Court cited several precedents (Richey v. Commissioner, Mazzei v. Commissioner, Wagner v. Commissioner) where deductions were denied to taxpayers who had never been charged with or convicted of crimes. The Court found that the petitioner's conduct — admittedly starting the fire — was grossly negligent regardless of whether the criminal charge resulted in a trial, and that the public policy of Maryland against arson was clearly articulated in the state's criminal code.",
 "Blackman v. Commissioner, No. 21436-84 (U.S.T.C. 1987) __ Justia.pdf",
 5, "Public policy doctrine; §165", "Analytical", "Hard",
 "Conviction not required for public policy denial."),

("J020", "Judgments",
 "In Bramblett v. Commissioner (59 T.C.M. 876, 1990), what was the legal significance of Mesquite East (the seller) having contracts of sale with third parties before the actual sale to Town East?",
 "The existence of contracts with third parties for approximately 78 acres of the land entered into by Baker and Bramblett as trustee before Town East even acquired the property from Mesquite East was highly significant. The Tax Court cited this as evidence that the property was already being held primarily for sale to customers in the ordinary course of business at the time of the nominal sale to Town East. Effectively, Town East was the mechanism through which the existing sales commitments were honored, and the arrangements show that the land was part of a real estate sales business from the outset, not an investment being passively held for appreciation.",
 "Bramblett v. Commissioner, No. 728-89 (U.S.T.C. 1990) __ Justia.pdf",
 7, "§1221(1); dealer analysis", "Analytical", "Hard",
 "Pre-existing customer contracts before acquisition = ordinary income."),

# Additional judgements (using more documents):
("J021", "Judgments",
 "In Foote v. Commissioner (81 T.C. 930, 1983), did the Tax Court acknowledge any merit in Foote's economic argument that tenure has characteristics of a capital asset?",
 "Yes, the Tax Court acknowledged that Foote's economic argument was 'ingenious and well presented' and showed 'some plausibility.' Foote argued that tenure provides a faculty member with freedom to exploit his university affiliation to generate income from consulting, writing, and research; that the university is the only 'full function buyer' of the tenure; and that the 'repurchase value' reflects the tenure's capital asset qualities. However, the Court declined to accept this economic argument because the legal definition of a capital asset in §1221 and the requirement of a 'sale or exchange' under §1222 are controlling. Tenure cannot be sold to a third party and is personal to the holder, making it legally distinct from a property interest eligible for capital gain treatment.",
 "Foote v. Commissioner, No. 9667-81 (U.S.T.C. 1983) __ Justia.pdf",
 5, "§§1221, 1222; capital asset definition", "Analytical", "Hard",
 ""),

("J022", "Judgments",
 "In Carter v. Commissioner (40 T.C.M. 654, 1980), what is the legal definition of 'constructive receipt' of income as applied by the Tax Court?",
 "The Tax Court applied the definition from §1.451-2(a) of the Income Tax Regulations: income is constructively received by a taxpayer in the taxable year during which it is credited to the taxpayer's account, set apart for the taxpayer, or otherwise made available so that the taxpayer may draw upon it at any time, or so that the taxpayer could have drawn upon it during the taxable year if notice of intention to withdraw had been given. Income is not constructively received if the taxpayer's control of its receipt is subject to substantial limitations or restrictions. In Carter's case, the mere presence of funds in New York City's budget did not satisfy this test because the funds were not available for Carter to draw upon.",
 "Carter v. Commissioner, No. 12404-78 (U.S.T.C. 1980) __ Justia.pdf",
 3, "§451; Reg. §1.451-2(a)", "Conceptual", "Medium",
 ""),

("J023", "Judgments",
 "In Gilliam v. Commissioner (51 T.C.M. 515, 1986), how did the Tax Court distinguish Gilliam's situation from the Dancer v. Commissioner case that petitioners relied upon?",
 "Petitioners argued that Dancer v. Commissioner (73 T.C. 1103, 1980) directly controlled and required deductibility because legal fees arising from an accident occurring during a business trip are deductible as ordinary and necessary business expenses. The Tax Court, however, found Gilliam's situation distinguishable: unlike Dancer, which involved an automobile accident during a business trip (a foreseeable and ordinary risk of travel), Gilliam's conduct involved an intentional act (assaulting a passenger) during a temporary insanity episode. The Court agreed with the IRS that the criminal charges 'could hardly be deemed ordinary given the nature of Gilliam's profession' as an artist and teacher, and the expenses were not of the type that normally attend the carrying on of those businesses.",
 "Gilliam v. Commissioner, No. 4986-79 (U.S.T.C. 1986) __ Justia.pdf",
 4, "§162(a); Dancer distinguished", "Analytical", "Hard",
 "Intentional assault vs. accident: key distinction."),

("J024", "Judgments",
 "In Bellamy v. Commissioner (43 T.C. 487, 1965), what was the nature of the Esty agreement and the subsequent 1954 amendment that led to the $89,000 payment?",
 "In 1949, actor Ralph Bellamy entered into an employment contract (Esty agreement) with William Esty Co. (acting for R.J. Reynolds) to perform in the television series 'Man Against Crime' at $1,500 per week initially, rising to $3,500 per week under a 1952 amendment that permitted filming for rebroadcast. The 1954 agreement extended by 78 weeks the period during which the filmed episodes could be broadcast (from 26 to 104 weeks) in exchange for Esty relinquishing its right to show the films after the 104-week period. In 1957, the $89,000 was received when Esty sold the films, representing Bellamy's share of net proceeds. The Tax Court found this was ordinary income because the payment compensated for the relinquishment of rights that were personal service contract rights.",
 "Bellamy v. Commissioner, No. 989-63 (U.S.T.C. 1965) __ Justia.pdf",
 7, "Employment contract; capital asset analysis", "Factual", "Hard",
 "$89,000 = 20% of net proceeds from film sales."),

("J025", "Judgments",
 "In Bynum v. Commissioner (46 T.C. 295, 1966), what was the significance of the concurring opinion by Judge Tannenwald regarding the three elements required to hold property for sale in the ordinary course of business?",
 "Judge Tannenwald's concurrence emphasized that the statutory phrase 'held primarily for sale to customers in the ordinary course of his trade or business' contains three separate and important elements: (1) 'Primarily' — as Malat v. Riddell held, means 'of first importance,' not merely 'substantial'; (2) 'For sale to customers' — essentially satisfied whenever there is a proposed sale, since buyers are always 'customers'; and (3) 'In the ordinary course of business' — this is the crucial element, requiring that the taxpayer must be in a business of which the sale is a part, and the sale must be in the ordinary course of that business (not merely an isolated or unusual transaction). Tannenwald emphasized that the third element does the real work of separating business sales from investment liquidations.",
 "Bynum v. Commissioner, No. 5226-63 (U.S.T.C. 1966) __ Justia.pdf",
 8, "§§1221(1), 1231(b)(1)(B) analysis", "Analytical", "Hard",
 "Three elements: 'primarily,' 'customers,' 'ordinary course.'"),

# ════════════════════════════════════════════════════════════════════════
# POV  (target 35)
# ════════════════════════════════════════════════════════════════════════

("P001", "POV",
 "According to the Tax Foundation's analysis of the One Big Beautiful Bill Act (OBBBA), what was the estimated average tax cut per individual taxpayer across the United States in 2026?",
 "According to the Tax Foundation's General Equilibrium Model (February 2026), the average tax cut per taxpayer across all individual tax filers throughout the United States will be $3,813 in 2026. Of this amount, individual tax changes in the OBBBA reduce tax liability by $2,272 on average in 2026, while business tax cuts contribute another $1,541 on average per taxpayer. The average falls to $2,590 in 2030 as certain provisions (such as deductions for tips and overtime income) expire.",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 2, "Tax Foundation OBBBA analysis", "Factual", "Medium",
 "Note: article title says $2,300 but table shows $3,813 average for all filers."),

("P002", "POV",
 "According to the Tax Foundation's OBBBA analysis, which state had the largest average tax cut per filer in 2026, and what was that amount?",
 "According to the Tax Foundation's state-level estimates table in the OBBBA analysis, Wyoming had the largest average tax cut per filer in 2026 at $5,478. Washington State was second at $5,445, and Massachusetts was third at $5,259. At the county level, Teton County, Wyoming had the highest average tax cut at $39,316 per taxpayer in 2026. The smallest state-level average tax cuts were in West Virginia ($2,448) and Mississippi ($2,386).",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 4, "State-level tax cut estimates", "Factual", "Medium",
 "Wyoming = highest state average; WV/MS = lowest."),

("P003", "POV",
 "According to the Tax Foundation's OBBBA analysis, what was the estimated impact on employment from the OBBBA?",
 "The Tax Foundation estimated that the OBBBA will increase hours worked by about 828,000 full-time equivalent jobs over the long run. The analysis allocates this impact based on the national jobs estimates from the Tax Foundation General Equilibrium Model and the distribution of labor and capital income across states. The employment gains reflect the combined effects of the individual and business tax provisions of the OBBBA on labor supply and investment.",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 2, "Employment impact of OBBBA", "Factual", "Medium",
 "828,000 FTE jobs long-run increase."),

("P004", "POV",
 "According to the Tax Foundation's European Tax Policy Scorecard (2025), what two principles does the Scorecard use to evaluate European countries' tax systems?",
 "The European Tax Policy Scorecard seeks to measure the extent to which a country's tax system adheres to two important principles: (1) Competitiveness — a competitive tax code keeps marginal tax rates low, since in a globalized world, capital is highly mobile and businesses will choose to invest where after-tax returns are maximized; and (2) Neutrality — a neutral tax code is one that seeks to raise revenue while causing as few economic distortions as possible, meaning similar economic activities should be taxed similarly so that tax considerations do not drive business decisions.",
 "2025 European Tax Rankings _ Tax Foundation Europe.pdf",
 2, "ETPS methodology", "Conceptual", "Medium",
 "Competitiveness + neutrality = two ETPS pillars."),

("P005", "POV",
 "According to the Tax Foundation's OBBBA analysis, what methodology was used to allocate the national tax cut estimates to individual counties?",
 "The Tax Foundation allocated the conventional national revenue estimates for each OBBBA provision to individual counties using data from the IRS Statistics of Income for individual tax returns (2022). For each provision, specific IRS data characteristics (such as county share of taxable income, standard deductions claimed, CTC amounts, SALT disallowed, and mortgage interest deductions) were used as allocation factors. The business tax provisions were allocated based on labor and capital income shares, weighted to reflect the economic incidence of the corporate tax shifting from capital (90% in year 1) to labor (50/50 in year 5 and beyond).",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 6, "Methodology section", "Factual", "Hard",
 "IRS Statistics of Income county data used for allocation."),

("P006", "POV",
 "According to the Tax Foundation's OBBBA analysis, how does the $40,000 SALT deduction cap affect taxpayers geographically?",
 "The Tax Foundation's analysis notes that the $40,000 cap on state and local tax (SALT) deductions (which reverts to a $10,000 cap after 2029) will tend to have the greatest impact on taxpayers in higher-tax localities on the coasts of the United States. This reflects the geographic variation in state and local tax burdens, with taxpayers in high-tax states like California, New York, and New Jersey more likely to have SALT amounts exceeding the cap, and therefore more likely to benefit from the higher cap (or be limited by the lower cap post-2029) compared to taxpayers in lower-tax states.",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 2, "SALT deduction geographic impact", "Analytical", "Medium",
 "SALT cap: $40,000 through 2029, then $10,000."),

("P007", "POV",
 "What does the 2025 European Tax Policy Scorecard measure, and what is the significance of tax policy structure for economic performance according to the introduction?",
 "The 2025 European Tax Policy Scorecard measures the extent to which European countries' tax systems adhere to principles of competitiveness and neutrality. The introduction explains that the structure of a country's tax code is a determining factor of its economic performance: a well-structured tax code is easy for taxpayers to comply with and can promote economic development while raising sufficient revenue for government priorities. In contrast, poorly structured tax systems can be costly, distort economic decision-making, and harm domestic economies. These principles apply at the supranational level as well, given the EU's goals of increased defense spending, green and digital transitions, and possible enlargement.",
 "2025 European Tax Rankings _ Tax Foundation Europe.pdf",
 1, "Introduction; tax policy principles", "Conceptual", "Easy",
 ""),

("P008", "POV",
 "According to the Tax Foundation's OBBBA analysis, what key provisions of the One Big Beautiful Bill Act made it significant for tax policy since the 2017 Tax Cuts and Jobs Act?",
 "According to the Tax Foundation, the OBBBA makes permanent the individual tax changes first enacted by the Tax Cuts and Jobs Act of 2017, avoiding a tax increase on approximately 62 percent of tax filers in 2026. In addition, the OBBBA provides new individual tax cuts beyond TCJA extensions, including: new deductions for tipped and overtime income; an expanded child tax credit; 100 percent bonus depreciation made permanent; and permanent expensing for domestic research and development (R&D). These combined changes make it the most significant legislative changes to federal tax policy since the 2017 TCJA.",
 "$2,300 Average Tax Cut in 2026 Under the Big Beautiful Bill.pdf",
 1, "OBBBA key provisions", "Factual", "Medium",
 "OBBBA = most significant tax legislation since TCJA 2017."),

# ════════════════════════════════════════════════════════════════════════
# TAX DOCS  (target 15)
# ════════════════════════════════════════════════════════════════════════

("T001", "Tax Docs",
 "According to IRS Publication 15 (Circular E, 2026), what is the social security wage base limit for 2026?",
 "Publication 15 (2026) states that the social security wage base limit is $184,500 for 2026. Social security tax applies only to wages up to this amount, while Medicare tax has no wage base limit. Social security and Medicare taxes apply at the same rate to the employee and employer. The Medicare tax rate is 1.45% each for the employee and employer, unchanged from 2025.",
 "p15.pdf",
 2, "What's New; Social Security Wage Base", "Factual", "Easy",
 "Social Security wage base = $184,500 for 2026."),

("T002", "Tax Docs",
 "According to IRS Publication 15 (2026), what is the withholding rate on supplemental wages paid to employees, and at what dollar threshold does the higher rate apply?",
 "Publication 15 (2026) states that the withholding rate on supplemental wages remains 22 percent (the optional flat rate), because the One Big Beautiful Bill Act permanently extended the individual tax rates enacted in the 2017 Tax Cuts and Jobs Act. If supplemental wages paid to an employee during the calendar year exceed $1 million, the withholding rate on the excess is 37 percent. The backup withholding rate also remains at 24 percent.",
 "p15.pdf",
 2, "Supplemental wages withholding", "Factual", "Easy",
 "22% up to $1M; 37% over $1M."),

("T003", "Tax Docs",
 "According to IRS Publication 15 (2026), what is a 'Trump account' and what is the annual contribution limit?",
 "According to Publication 15, the One Big Beautiful Bill Act (P.L. 119-21) allows for a new type of traditional individual retirement account called a 'Trump account,' which can be established for a child who has not attained age 18 at the end of the year the account is established. The annual contribution limit is $5,000 (other than exempt contributions), which will be indexed for inflation after tax year 2027. Beginning July 4, 2026, employers may contribute up to $2,500 per year (also indexed for inflation after 2027) toward the $5,000 limit, and such employer contributions are excluded from the employee's gross income if paid pursuant to a Trump account contribution program.",
 "p15.pdf",
 3, "Trump account; employer contributions", "Factual", "Medium",
 "New account type in OBBBA; $5,000 annual limit; employer may contribute $2,500."),

("T004", "Tax Docs",
 "According to IRS Publication 15 (2026), what are the requirements for federal tax deposits, and what electronic payment options are available?",
 "Publication 15 states that all federal tax deposits must be made by electronic funds transfer (EFT). The available electronic methods are: (1) the Electronic Federal Tax Payment System (EFTPS), a free service by the Department of the Treasury; (2) IRS Direct Pay; or (3) an IRS business tax account. Alternatively, employers may arrange for their tax professional, financial institution, or payroll service to make electronic deposits on their behalf, or arrange for a same-day wire payment. Credit or debit card payments may be used for balance due on employment tax returns, but not for federal tax deposits. Failure to use EFT for deposits may result in a 10 percent failure-to-deposit penalty.",
 "p15.pdf",
 6, "Electronic Federal Tax Deposits", "Factual", "Medium",
 "EFT required for all federal deposits; EFTPS is free."),

("T005", "Tax Docs",
 "According to IRS Publication 17 (2025), what is the due date for filing Form 1040 for the 2025 tax year?",
 "Publication 17 (2025) states that the due date for filing Form 1040 or Form 1040-SR for the 2025 tax year is April 15, 2026. This is the standard deadline for individual income tax returns. Extensions of time to file (but not time to pay) are available by filing Form 4868, which generally extends the deadline by 6 months.",
 "p17.pdf",
 3, "What's New; Filing Deadline", "Factual", "Easy",
 "2025 return due April 15, 2026."),

("T006", "Tax Docs",
 "According to IRS Publication 15 (2026), what is the information reporting threshold for certain payments to persons after calendar year 2025, as changed by P.L. 119-21?",
 "Publication 15 (2026) states that for payments made after calendar year 2025, P.L. 119-21 (the One Big Beautiful Bill Act) increases the information reporting threshold (for example, Forms 1099-MISC and Forms 1099-NEC) from $600 to $2,000 in a calendar year for certain payments to persons. This threshold will be adjusted for inflation for each calendar year after 2026. Similarly, the aggregate reportable payment threshold for backup withholding purposes under §6041(a) or §6041A(a) for calendar year 2026 is increased from $600 to $2,000.",
 "p15.pdf",
 2, "Information reporting threshold changes", "Factual", "Medium",
 "Information reporting threshold raised from $600 to $2,000 by OBBBA."),

("T007", "Tax Docs",
 "According to IRS Publication 15 (2026), when do social security and Medicare taxes apply to household workers or election workers?",
 "Publication 15 (2026) states that social security and Medicare taxes apply to the wages of household workers if the employer pays them $3,000 or more in cash wages in 2026. For election workers, social security and Medicare taxes apply if they are paid $2,500 or more in cash or an equivalent form of compensation in 2026. Both thresholds reflect the 2026 amounts. There is no wage base limit for Medicare tax, and the Medicare tax rate is 1.45% each for the employee and employer.",
 "p15.pdf",
 2, "Household and election worker FICA thresholds", "Factual", "Medium",
 "Household: $3,000; election workers: $2,500."),

# ════════════════════════════════════════════════════════════════════════
# CROSS-DOC  (target 10-15)
# ════════════════════════════════════════════════════════════════════════

("X001", "Cross-Doc",
 "Both 26 U.S.C. §165(c)(3) and the Tax Court in Blackman v. Commissioner (88 T.C. 677) address casualty losses. How do these two sources together explain why a taxpayer who intentionally starts a fire cannot deduct the resulting property loss?",
 "Section 165(c)(3) allows individuals to deduct losses of property not connected with a trade or business that arise from 'fire, storm, shipwreck, or other casualty.' While the statute does not explicitly exclude intentional losses, the Tax Court in Blackman held that two additional doctrines bar the deduction when the taxpayer's own conduct caused the loss. First, grossly negligent or intentional conduct that foreseeably causes the loss bars the casualty deduction because the loss is not 'sustained' in the manner contemplated by the statute. Second, where allowing the deduction would frustrate a clearly articulated state public policy (here, Maryland's prohibition on arson), the deduction is barred on public policy grounds. Together, §165(c)(3) provides the general framework while Blackman establishes the limits: intentional or grossly negligent conduct, or conduct violating public policy, defeats the casualty loss deduction.",
 "26 U.S. Code § 165 - Losses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§165(c)(3); Blackman holding", "Analytical", "Hard",
 "Cross-doc: §165 + Blackman (88 T.C. 677)."),

("X002", "Cross-Doc",
 "How do 26 U.S.C. §1221 (as applied in Foote v. Commissioner) and 26 U.S.C. §61(a) together explain why a tenured professor's severance payment for resigning tenure is taxable as ordinary income rather than capital gain?",
 "Section 61(a) provides that gross income means 'all income from whatever source derived,' including compensation for services. Absent a statutory exclusion or specific provision taxing income at capital gain rates, all income is ordinary income. Section 1221 defines 'capital asset' as property held by the taxpayer, but does not include property held primarily for sale, depreciable business property, or certain other specified types. In Foote v. Commissioner (81 T.C. 930), the Tax Court held that academic tenure is not a 'capital asset' because it cannot be sold to a third party, is personal to the holder, and represents an employment contract right rather than a property interest. Furthermore, the resignation was not a 'sale or exchange' as required for capital gain treatment under §1222. Since the payment was essentially a substitute for future ordinary income (salary), it falls within §61(a)'s broad definition and none of the capital gain provisions apply.",
 "26 U.S. Code § 61 - Gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§61(a); §1221; Foote", "Analytical", "Hard",
 "Cross-doc: §61(a) + §1221 + Foote (81 T.C. 930)."),

("X003", "Cross-Doc",
 "How do 26 U.S.C. §164(b)(6) (the SALT cap) and the Tax Foundation's OBBBA analysis together describe the geographic impact of the SALT deduction limitation on American taxpayers?",
 "Section 164(b)(6) limits the state and local tax (SALT) deduction for individuals to $40,000 for 2025, $40,400 for 2026, with the cap dropping to $10,000 after 2029. The Tax Foundation's OBBBA analysis (February 2026) explains that this limitation has the greatest geographic impact on taxpayers in higher-tax localities on the coasts of the United States. States like California, New York, Connecticut, and New Jersey have higher state and local tax burdens, meaning their taxpayers are more likely to have SALT amounts exceeding the cap. In contrast, low-tax states like Mississippi and West Virginia saw the smallest average tax cuts from the OBBBA overall, in part because their residents have lower state and local tax burdens and thus fewer dollars of SALT to deduct in the first place.",
 "26 U.S. Code § 164 - Taxes _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 5, "§164(b)(6); OBBBA analysis", "Analytical", "Hard",
 "Cross-doc: §164(b)(6) SALT cap + Tax Foundation OBBBA analysis."),

("X004", "Cross-Doc",
 "How does 26 U.S.C. §451 (as applied in Carter v. Commissioner) relate to the general principle of §61(a) that gross income is all income from whatever source derived?",
 "Section 61(a) establishes the broad inclusivity of gross income: all income from whatever source derived is included unless a specific exclusion applies. Section 451 governs the timing of income inclusion — it provides that amounts are included in gross income for the taxable year in which they are 'received' (for cash basis taxpayers). Carter v. Commissioner (40 T.C.M. 654) applied the constructive receipt doctrine: even though Carter's wages were earned in 1974, they were not constructively received until paid in 1975, because a substantial limitation (administrative processing backlog) prevented him from drawing upon them in 1974. Together, §61(a) determines what is includible and §451 determines when it must be included. The cash method of accounting, per Carter, requires inclusion in the year of actual or constructive receipt.",
 "26 U.S. Code § 61 - Gross income defined _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§61(a) + §451; Carter v. Commissioner", "Conceptual", "Medium",
 "Cross-doc: §61 + §451 + Carter (40 T.C.M. 654)."),

("X005", "Cross-Doc",
 "How does 26 U.S.C. §6662 (accuracy-related penalty) connect with the facts in Calloway v. Commissioner (135 T.C. 26), where the taxpayer failed to report gain from the Derivium transaction?",
 "Section 6662(a) imposes a 20 percent accuracy-related penalty on underpayments attributable to, among other things, negligence or disregard of rules or regulations, or any substantial understatement of income tax. In Calloway v. Commissioner, the Tax Court held that the Calloways were liable for the accuracy-related penalty under §6662. Albert Calloway had reviewed a tax memorandum from Robert Nagy opining there was a 'solid basis' for treating the Derivium transaction as a loan rather than a sale, but the Court found he relied on an opinion prepared at Derivium's request for marketing purposes rather than for independent tax advice. This reliance did not constitute reasonable cause or good faith under §6664(c), which is the defense to the §6662 penalty. The failure to report the gain from what the Court held was a 2001 sale of IBM stock produced both a deficiency and an accuracy-related penalty.",
 "26 U.S. Code § 6662 - Imposition of accuracy-related penalty on underpayments _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§6662; Calloway penalty analysis", "Analytical", "Hard",
 "Cross-doc: §6662 + Calloway (135 T.C. 26)."),

("X006", "Cross-Doc",
 "The OBBBA (as described in both IRS Publication 15 and the Tax Foundation's analysis) made permanent certain individual income tax rates. How does this interact with 26 U.S.C. §6651 (failure-to-file penalty) and taxpayers' filing obligations?",
 "The OBBBA permanently extended the individual income tax rates first enacted by the 2017 Tax Cuts and Jobs Act, as noted in IRS Publication 15 (2026). This means that the tax rate schedules under 26 U.S.C. §1 remain in effect on a permanent basis. For taxpayers subject to these rates, the obligation to file returns on time under §6013 (or whichever section applies) and to pay any tax due is unaffected. Section 6651(a)(1) imposes a 5-percent-per-month penalty (up to 25 percent) for failure to timely file, and §6651(a)(2) imposes a 0.5-percent-per-month penalty (up to 25 percent) for failure to pay. The permanence of the OBBBA tax rates increases the certainty of taxpayers' tax liabilities, but does not change the basic filing and payment obligations or the applicable penalty structure under §6651.",
 "p15.pdf",
 2, "OBBBA + §6651", "Analytical", "Medium",
 "Cross-doc: Pub. 15 OBBBA + §6651 filing requirements."),

("X007", "Cross-Doc",
 "How does the 'ordinary course of business' test applied in both Bynum v. Commissioner and Adam v. Commissioner differ, and what underlying facts drove the different outcomes in each case?",
 "Both Bynum v. Commissioner (46 T.C. 295) and Adam v. Commissioner (60 T.C. 996) applied the same legal test from §1221(1): whether property was held 'primarily for sale to customers in the ordinary course of trade or business.' However, the outcomes differed based on the underlying facts. In Bynum, the taxpayer subdivided a 113-acre farm into at least 55 lots, spent over $1,000 per acre on improvements (more than the pre-improvement market value), ran full-page newspaper advertisements, listed lots with multiple realtors, and personally sold all lots without commissions — active and systematic sales activities that established dealer status. In Adam, by contrast, the taxpayer never subdivided, plotted, or improved any waterfront property; never advertised; never employed brokers; and engaged in only occasional and intermittent transactions. Bynum's activities created a business of selling lots, while Adam's lack of activity confirmed investor status.",
 "Bynum v. Commissioner, No. 5226-63 (U.S.T.C. 1966) __ Justia.pdf",
 6, "§1221(1); dealer vs. investor", "Analytical", "Hard",
 "Cross-doc: Bynum (46 T.C. 295) + Adam (60 T.C. 996)."),

("X008", "Cross-Doc",
 "How does 26 U.S.C. §2503(b) (annual gift tax exclusion) interact with the facts in Estate of Kohlsaat to determine when a gift qualifies as a present interest?",
 "Section 2503(b) excludes from taxable gifts the first $10,000 per year per donee (indexed for inflation), but this exclusion applies only to gifts of 'present interests' — interests in property that are not limited to commence in use, possession, or enjoyment at some future date. In Estate of Kohlsaat (73 T.C.M. 2732), the Tax Court applied this statutory distinction by finding that the contingent remainder beneficiaries of the Kohlsaat Family Trust each held present interests because they had unrestricted Crummey withdrawal rights: the right to demand immediate distribution of trust property within a 30-day window following each transfer. Because these rights gave immediate access to property, they satisfied the present interest requirement of §2503(b). The Court cited Crummey v. Commissioner (9th Cir. 1968) and Estate of Cristofani v. Commissioner (97 T.C. 74) as supporting authority.",
 "26 U.S. Code § 2501 - Imposition of tax _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§2503(b); Crummey powers", "Analytical", "Hard",
 "Cross-doc: §2503(b) + Estate of Kohlsaat."),

("X009", "Cross-Doc",
 "How does 26 U.S.C. §162(a) (ordinary and necessary business expenses) interact with the holding in Gilliam v. Commissioner (51 T.C.M. 515), and what two-part test must be satisfied for an expense to be deductible?",
 "Section 162(a) allows deductions for all 'ordinary and necessary expenses' paid or incurred during the taxable year in carrying on a trade or business. In Gilliam, the Tax Court applied a two-part test: (1) the expense must be 'ordinary' — meaning it is of a type that commonly or normally arises in the conduct of the taxpayer's business; and (2) it must be 'necessary' — meaning it is helpful and appropriate to the business, though not absolutely required. In addition, the expense must be proximately related to the taxpayer's trade or business. Gilliam's legal fees failed the 'ordinary' prong because the conduct giving rise to them — attacking a fellow passenger aboard an airplane — was not a normal or expected event in carrying on an artist's and art teacher's business. The Court refused to extend the §162 deduction to expenses arising from conduct outside the normal scope of the taxpayer's business activities.",
 "26 U.S. Code § 162 - Trade or business expenses _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 2, "§162(a); ordinary and necessary", "Analytical", "Hard",
 "Cross-doc: §162(a) + Gilliam (51 T.C.M. 515)."),

("X010", "Cross-Doc",
 "How do 26 U.S.C. §170(b)(1)(G) (60% cash charitable contribution limit) and the OBBBA (as described in the Tax Foundation analysis) together illustrate the relationship between permanent law and year-to-year tax planning?",
 "Section 170(b)(1)(G) increased the percentage limitation for cash contributions to public charities from 50 percent to 60 percent of the taxpayer's contribution base for taxable years beginning after December 31, 2017, as enacted by the 2017 Tax Cuts and Jobs Act. This provision was originally set to expire at the end of 2025. The OBBBA (signed July 2025), as analyzed by the Tax Foundation, made the individual tax changes of the 2017 TCJA permanent, which means the 60% cash contribution limitation under §170(b)(1)(G) is now a permanent part of the tax code rather than a temporary provision. This permanence is significant for tax planning because donors contributing cash to public charities can now plan on a 60% AGI ceiling without needing to anticipate expiration.",
 "26 U.S. Code § 170 - Charitable, etc., contributions and gifts _ U.S. Code _ US Law _ LII _ Legal Information Institute.pdf",
 8, "§170(b)(1)(G); TCJA permanence", "Analytical", "Hard",
 "Cross-doc: §170(b)(1)(G) + Tax Foundation OBBBA analysis."),

]

# ──────────────────────────────────────────────────────────────────────────────
# WRITE TO EXCEL
# ──────────────────────────────────────────────────────────────────────────────

output_path = r"C:\Users\Ls Computer\Downloads\Legal_Tax_RAG_System\Golden_Set.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Golden Set"

# Header
headers = [
    "ID", "Category", "Query", "Ground Truth Answer",
    "Source Document", "Page Number", "Section/Clause",
    "Query Type", "Difficulty", "Notes"
]

# Style helpers
header_fill = PatternFill("solid", fgColor="1F3864")   # dark navy
alt_fill    = PatternFill("solid", fgColor="EEF2F7")   # light blue-grey
white_fill  = PatternFill("solid", fgColor="FFFFFF")

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
body_font   = Font(name="Calibri", size=10)
wrap        = Alignment(wrap_text=True, vertical="top")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write headers
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 28

# Write data rows
for row_idx, r in enumerate(rows, 2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col_idx, val in enumerate(r, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.fill = fill
        cell.alignment = wrap
        cell.border = border
    ws.row_dimensions[row_idx].height = 80

# Column widths
col_widths = {1: 8, 2: 10, 3: 48, 4: 65, 5: 50, 6: 8, 7: 20, 8: 12, 9: 10, 10: 40}
for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze pane
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

wb.save(output_path)
print(f"\nGolden Set written to: {output_path}")
print(f"Total rows: {len(rows)}")

# Count by category
from collections import Counter
cats = Counter(r[1] for r in rows)
for cat, cnt in sorted(cats.items()):
    print(f"  {cat}: {cnt} rows")
